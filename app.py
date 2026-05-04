from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from openpyxl import Workbook
import os, io, math, requests, zipfile, re

FULCRUM_SESSION = os.environ.get('FULCRUM_SESSION', '')
FULCRUM_BASE = 'https://integrasystems.fulcrumpro.com/api'

def fulcrum_headers():
    return {
        'Cookie': f'fp-auth-token-integrasystems={FULCRUM_SESSION}',
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }

def fulcrum_upload_headers():
    return {
        'Cookie': f'fp-auth-token-integrasystems={FULCRUM_SESSION}',
        'Accept': 'application/json'
    }

def get_existing_item(part_number):
    """Returns (fulcrum_id, []) or (None, []) if not found. Uses cached lookup table."""
    return item_id_cache.get(part_number.upper(), (None, []))

def build_item_id_cache(top_assembly_id):
    """Fetch assembly from Fulcrum and build part number -> ID cache."""
    if not FULCRUM_SESSION:
        return {}
    try:
        r = requests.get(f'{FULCRUM_BASE}/items/{top_assembly_id}',
            headers=fulcrum_headers(), timeout=10)
        print(f'Fulcrum assembly fetch: HTTP {r.status_code}')
        if r.status_code == 200:
            data = r.json()
            cache = {}
            cache[data.get('number','').upper()] = (data.get('id'), [])
            routing = data.get('routing', {})
            input_items = routing.get('inputItems', [])
            for item in input_items:
                ref = item.get('itemReference', {})
                pn = ref.get('number', '').upper()
                item_id = ref.get('id', '')
                if pn and item_id:
                    cache[pn] = (item_id, [])
            print(f'Built cache with {len(cache)} parts: {list(cache.keys())}')
            return cache
    except Exception as e:
        print(f'Fulcrum cache build error: {e}')
    return {}


def extract_attachments_from_zip(zip_bytes):
    attachments = {}
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            for name in z.namelist():
                basename = os.path.basename(name)
                if not basename: continue
                ext = os.path.splitext(basename)[1].lower()
                if ext not in ['.pdf', '.sldprt', '.sldasm', '.step', '.stp']:
                    continue
                stem = os.path.splitext(basename)[0].upper()
                pn = re.sub(r'-[Rr]?[0-9]{1,2}$', '', stem)
                if pn not in attachments:
                    attachments[pn] = []
                attachments[pn].append((basename, z.read(name)))
    except Exception as e:
        print(f'Zip extraction error: {e}')
    return attachments

def find_xlsx_in_zip(zip_bytes):
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            for name in z.namelist():
                basename = os.path.basename(name)
                if basename.endswith('.xlsx') and not basename.startswith('~'):
                    return basename, z.read(name)
    except Exception as e:
        print(f'Zip xlsx error: {e}')
    return None, None

def build_attachment_zip(assembly_number, attachments, parts):
    buf = io.BytesIO()
    part_numbers = set(p['pn'].upper() for p in parts)
    part_numbers.add(assembly_number.upper())
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        for pn, files in attachments.items():
            if any(pn.startswith(p) or p.startswith(pn) or pn == p for p in part_numbers):
                for filename, data in files:
                    z.writestr(f'{assembly_number}/{filename}', data)
    buf.seek(0)
    return buf

SPEED_TABLE = {
    'steel': {0.5:125,0.75:125,0.9:125,1.0:125,1.1:125,1.2:125,1.5:125,1.6:125,2.0:90,2.5:32,3.0:32,4.0:25,5.0:21,6.0:18,8.0:14,10.0:11},
    'aluminium': {0.5:125,0.75:125,0.9:125,1.0:125,1.1:125,1.2:125,1.5:125,1.6:125,2.0:85,2.5:35,3.0:35,4.0:17.5,5.0:9,8.0:4},
    'stainless': {0.5:125,0.75:125,0.9:125,1.0:125,1.1:125,1.2:125,1.5:125,2.0:70,2.5:42.5,3.0:42.5,5.0:15},
    'brass': {0.5:90,0.75:90,0.9:90,1.0:90,1.1:90,1.6:90},
}
MATERIAL_CATEGORY = {
    'mild steel':'steel','zinc anneal':'steel','galvanised':'steel','galvanized':'steel','steel':'steel',
    'aluminium':'aluminium','aluminum':'aluminium','stainless steel':'stainless',
    '18-8 stainless steel':'stainless','316 stainless steel':'stainless','s/s 304 2b':'stainless','brass':'brass',
}

PC_COLOURS = [
    'PC- Rapidcure Black Matt','PC-Black Scylla','PC-Blaze Blue','PC-Blaze Blue Gloss',
    'PC-Classic Pearl White Gloss','PC-Colourbond Monument','PC-Colourbond Surfmist',
    'PC-Dulux Bright White','PC-Duralloy Pearl White Gloss','PC-Duratec Black',
    'PC-Duratec Intensity Storm Satin','PC-ELEMENTS2 BLACK NIGHTSKY FLAT','PC-Evergreen',
    'PC-Green Mistletoe Gloss','PC-Intensity Flame Red','PC-Lemon Yellow',
    'PC-MA494A INTERPON 610 WHITE GLOSS','PC-Olde Pewter','PC-Orange X15','PC-Oyster Matt',
    'PC-Palladium Silver','PC-Protexture Black Flat','PC-Protexture Silver Pearl',
    'PC-RAL 7047 TELEGREY 4 SATIN','PC-Ripple Graphite','PC-Safety Yellow',
    'PC-Sahara Ebony Black','PC-Shale Grey GL284A Interpon','PC-Signal Red','PC-Special White',
    'PC-Telegray 4 RAL 7047','PC-Textura Black','PC-Textura White','PC-Trim Black',
    'PC-Trim Black replacement'
]

COLOUR_MAP = {c.lower().replace('pc-','').replace('pc- ','').strip(): c for c in PC_COLOURS}

def match_colour(raw_colour):
    if not raw_colour: return None
    raw = str(raw_colour).strip().lower()
    for key, pc in COLOUR_MAP.items():
        if key in raw or raw in key:
            return pc
    return None

def calc_powder(x, y):
    if not x or not y: return None, None
    area = (x/1000) * (y/1000) * 2
    powder_kg = area / 5
    if x * y <= 62500:
        time_min = 0.5
    else:
        time_min = 7 * (x/1000) * (y/1000)
    return round(powder_kg, 4), round(time_min, 2)

def get_cutting_speed(material, thickness_str):
    if not material or not thickness_str: return None
    cat = MATERIAL_CATEGORY.get(material.strip().lower())
    if not cat: return None
    try: thick = float(str(thickness_str).strip())
    except: return None
    table = SPEED_TABLE[cat]
    if thick in table: return table[thick]
    return table[min(table.keys(), key=lambda t: abs(t-thick))]

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

SETUP_SECS = 600
DEF_SPB = 30

FULCRUM_OP_MAP = {
    'laser cutting':'Laser Cutting','lasercut':'Laser Cutting','laser cut':'Laser Cutting',
    'press brake bending':'Press Brake Bending','press brake':'Press Brake Bending',
    'panel bending':'Panel Bending',
    'laser welding':'Laser Welding',
    'powdercoating':'Powder Coating','powder coating':'Powder Coating','powdercoat':'Powder Coating',
    'assembly':'Assembly','clinching':'Clinching','3d printing':'3D Printing',
    'outside processing':'Outsourced Fabrication','outsourced fabrication':'Outsourced Fabrication',
}

JUNK_MATS = ['not specified','material <','n/a','raw']

ITEM_COLS = ['Item Rev','Number','Description','RevisionNumber','GL Code','GL Code 2',
    'ItemOrigin','MinimumStockOnHand','MinimumProductionQuantity','Tags','UnitOfMeasureName',
    'LotTracking','IsSellItem','IsTaxable','BuildToStock','AllowContinuousFlow','Category',
    'Shape.Name','Material','Grade','Thickness','DimensionUnitOfMeasure','Length','Width',
    'Height','Notes','Shipping.Class','Shipping.NMFC','Shipping.IsHazMat','Shipping.UnitWeight',
    'Shipping.TariffCode','Shipping.CountryOfOrigin','Default Location','DimensionUnitOfMeasure',
    'Height','IsDraftItem','IsArchived','UsePartialPieceTracking']

BOM_H = ['ParentNumber','ParentRevisionNumber','ChildNumber','ChildRevisionNumber','Units Required','Nest Width','Nest Length','Units Created','RoutingOperation']
ROUT_H = ['Number','RevisionNumber','Operation','Order','Equipment','Instructions',
    'Setup Time Type','Setup Time','Setup Time Unit','Labor Time Type','Labor Time','Labor Time Unit',
    'Machine Tracking?','Machine Time Type','Machine Time','Machine Time Unit','Lead Days','Vendor','Cost Option','Cost']

EMPTY_SHEETS = {
    'Material Items':['Material Item Name','Material Name','Material','Grade','Finish','Form','Dimension','Length','Width'],
    'Vendor Item Details':['Number','RevisionNumber','Description','Vendor.Name','Vendor Item Name','Vendor Item Description','VendorLeadTimeInDays','PurchasingLink','IsPrimary','MinimumOrderQuantity','NotesToVendor','Vendor Unit Of Measure','VendorUnitOfMeasure.VendorQuantity','VendorUnitOfMeasure.FulcrumQuantity','Internal UoM'],
    'Customer Item Details':['Item Rev','Number','RevisionNumber','Standard Description','Customer External System ID','Customer Name','Customer Item Number','Customer Item Description','Customer Item Price','Unit of Measure'],
    'Price Breaks':['Item Rev','Number','RevisionNumber','Description','Break Quantity','Customer','Customer Item Number','Customer Item Description','Vendor','Vendor Item Number','Vendor Item Description','Customer Tier Name','Unit Price','Price Break Type'],
    'Inventory':['Item Rev','Number','RevisionNumber','Description','Location','Lot','Quantity','Unit of Measure','Material value','Labor value','Outside Processing value','Machine value'],
    'Sales UOMs':['Number','RevisionNumber','Custom','External Unit Of Measure','Internal Quantity','External Quantity'],
    'Item Inventory':['Number','RevisionNumber','Item_Inventory'],
}
ALL_SHEETS = ['Directions','Reference Data','Items','Material Items',
    'Vendor Item Details','Customer Item Details','Bill of Materials',
    'Routing','Price Breaks','Inventory','Sales UOMs','Item Inventory']

sessions = {}
item_id_cache = {}

def clean(v):
    if v is None: return ''
    return str(v).replace('\n',' ').replace('\r',' ').strip()

def is_junk_mat(m):
    if not m: return True
    v = m.lower()
    return any(j in v for j in JUNK_MATS) or not v.strip()

def is_numeric(v):
    if v is None: return False
    try: float(str(v).strip()); return True
    except: return False

def norm_proc(raw):
    if not raw: return None
    s = str(raw).strip().lstrip('-').strip().lower().replace('\n',' ').replace('\r',' ').replace('  ',' ')
    return FULCRUM_OP_MAP.get(s)

def find_col(header_row, *names):
    for i, val in enumerate(header_row):
        if val is None: continue
        v = str(val).strip().lower().replace('\n',' ').replace('  ',' ')
        for name in names:
            n = name.lower()
            # Exact match for short names, partial for longer ones
            if len(n) <= 2:
                if v == n:
                    return i
            else:
                if n in v:
                    return i
    return None

def parse_bom(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Detect columns by header name
    header = rows[0]
    col_x     = find_col(header, 'x') or 14
    col_y     = find_col(header, 'y') or 15
    col_bends = find_col(header, 'bend') or 16
    col_outer = find_col(header, 'outer') or 17
    col_inner = find_col(header, 'inner') or 18

    parts = []
    labor = {}
    pending_pn = None

    for i, row in enumerate(rows):
        if i == 0: continue

        pn_raw = row[1]
        pn = str(pn_raw).strip() if pn_raw else ''

        if not pn and pending_pn:
            x_val = row[col_x] if col_x < len(row) else None
            if x_val is not None:
                x = x_val
                y = row[col_y] if col_y < len(row) else None
                bends = float(row[col_bends]) if col_bends < len(row) and is_numeric(row[col_bends]) else None
                outer = float(row[col_outer]) if col_outer < len(row) and is_numeric(row[col_outer]) else None
                inner = float(row[col_inner]) if col_inner < len(row) and is_numeric(row[col_inner]) else None
            labor[pending_pn]['x'] = x
            labor[pending_pn]['y'] = y
            labor[pending_pn]['bends'] = bends
            labor[pending_pn]['outer'] = outer
            labor[pending_pn]['inner'] = inner
            p = next((p for p in parts if p['pn'] == pending_pn), None)
            if p and not labor[pending_pn]['speed']:
                labor[pending_pn]['speed'] = get_cutting_speed(p['mat'], p['thick'])
            continue

        if not pn: continue

        item_no = str(row[0]).strip() if row[0] else ''
        desc = clean(row[2])
        mat_r = clean(row[3])
        mat = '' if is_junk_mat(mat_r) else mat_r
        thick = clean(row[4]) if is_numeric(row[4]) else ''
        colour_raw = clean(row[5])
        colour = match_colour(colour_raw)
        try: qty = int(float(str(row[13]))) if row[13] else 1
        except: qty = 1

        procs = []
        for col in range(7, 13):
            op = norm_proc(row[col])
            if op and op not in procs:
                procs.append(op)

        indent = item_no.count('.')
        has_procs = bool(procs)
        is_make = has_procs or (mat and not is_junk_mat(mat))

        parts.append({'pn':pn,'indent':indent,'desc':desc,'mat':mat,'thick':thick,
                      'qty':qty,'processes':procs,'origin':'Make' if is_make else 'Buy',
                      'colour':colour,'item_no':item_no})
        labor[pn] = {'outer':None,'inner':None,'speed':None,'bends':None,'spb':DEF_SPB,'x':None,'y':None}
        pending_pn = pn

    return parts, labor

def calc_labor(pn, op, labor):
    d = labor.get(pn, {})
    if op == 'Laser Cutting':
        o = d.get('outer') or 0
        i = d.get('inner') or 0
        s = d.get('speed')
        if s and s > 0: return math.ceil((o+i)/s)
    if op in ('Press Brake Bending','Panel Bending'):
        b = d.get('bends')
        s = d.get('spb') or DEF_SPB
        if b and b > 0: return math.ceil(b*s)
    if op == 'Powder Coating':
        x = d.get('x'); y = d.get('y')
        if x and y:
            _, time_min = calc_powder(x, y)
            if time_min: return math.ceil(time_min * 60)
    return None

def irow(num, desc, origin, mat='', thick=''):
    r = [''] * len(ITEM_COLS)
    r[1]=num; r[2]=desc; r[6]=origin; r[10]='Piece'
    r[18]=mat; r[20]=thick; r[35]='FALSE'
    return r

def pc_irow(pc_name):
    r = [''] * len(ITEM_COLS)
    r[1]=pc_name; r[2]=pc_name; r[6]='Buy'; r[10]='Piece'; r[35]='FALSE'
    return r

def build_workbook(item_data, bom_data, rout_data):
    wb = Workbook()
    wb.remove(wb.active)
    for name in ALL_SHEETS:
        ws = wb.create_sheet(name)
        if name == 'Items':
            for row in item_data: ws.append(row)
        elif name == 'Bill of Materials':
            for row in bom_data: ws.append(row)
        elif name == 'Routing':
            for row in rout_data: ws.append(row)
        elif name in EMPTY_SHEETS:
            ws.append(EMPTY_SHEETS[name])
    return wb

def build_hierarchy(parts, top):
    seen = {}; bom_rows = []; stack = {0: top}
    for p in parts:
        if p['pn'] not in seen: seen[p['pn']] = p
        parent = stack.get(p['indent']-1, top) if p['indent'] > 0 else top
        stack[p['indent']] = p['pn']
        for k in list(stack.keys()):
            if k > p['indent']: del stack[k]
        bom_rows.append({'parent':parent,'child':p['pn'],'qty':p['qty']})
    return seen, bom_rows

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error':'No file'}), 400
    f = request.files['file']
    filename = f.filename
    file_bytes = f.read()

    attachments = {}
    if filename.lower().endswith('.zip'):
        attachments = extract_attachments_from_zip(file_bytes)
        bom_filename, bom_bytes = find_xlsx_in_zip(file_bytes)
        if not bom_bytes:
            return jsonify({'error':'No xlsx found inside zip'}), 400
        filename = bom_filename
        file_bytes = bom_bytes

    top = os.path.splitext(filename.split('_')[0].strip())[0]

    try:
        parts, labor = parse_bom(file_bytes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    session_id = top
    sessions[session_id] = {'parts':parts,'labor':labor,'top':top,'filename':filename,'attachments':attachments}

    laser = [p for p in parts if 'Laser Cutting' in p['processes']]
    press = [p for p in parts if 'Press Brake Bending' in p['processes']]
    panel = [p for p in parts if 'Panel Bending' in p['processes']]
    powder = [p for p in parts if 'Powder Coating' in p['processes']]

    return jsonify({
        'session_id': session_id, 'top': top, 'total': len(parts)+1, 'has_attachments': bool(attachments),
        'laser': [{'pn':p['pn'],'desc':p['desc'],'mat':p['mat'],'thick':p['thick'],
                   'outer':labor[p['pn']]['outer'],'inner':labor[p['pn']]['inner'],
                   'speed':labor[p['pn']]['speed'],'x':labor[p['pn']]['x'],'y':labor[p['pn']]['y']} for p in laser],
        'press': [{'pn':p['pn'],'desc':p['desc'],'mat':p['mat'],'thick':p['thick'],
                   'bends':labor[p['pn']]['bends'],'spb':labor[p['pn']]['spb']} for p in press],
        'panel': [{'pn':p['pn'],'desc':p['desc'],'mat':p['mat'],'thick':p['thick'],
                   'bends':labor[p['pn']]['bends'],'spb':labor[p['pn']]['spb']} for p in panel],
        'powder': [{'pn':p['pn'],'desc':p['desc'],'colour':p['colour'],
                    'x':labor[p['pn']]['x'],'y':labor[p['pn']]['y']} for p in powder],
    })

@app.route('/update_labor', methods=['POST'])
def update_labor():
    data = request.json
    sid = data.get('session_id')
    if sid not in sessions: return jsonify({'error':'Session not found'}), 404
    pn = data['pn']; field = data['field']; value = data['value']
    try: value = float(value)
    except: value = None
    sess = sessions[sid]
    if pn in sess['labor']: sess['labor'][pn][field] = value
    result = {}
    for p in sess['parts']:
        if p['pn'] == pn:
            for op in p['processes']:
                result[op] = calc_labor(pn, op, sess['labor'])
    return jsonify({'pn':pn,'labor':result})

@app.route('/download/attachments/<session_id>')
def download_attachments(session_id):
    if session_id not in sessions: return jsonify({'error':'Session not found'}), 404
    sess = sessions[session_id]
    attachments = sess.get('attachments', {})
    parts = sess['parts']
    top = sess['top']
    if not attachments:
        return jsonify({'error':'No attachments found'}), 404
    buf = build_attachment_zip(top, attachments, parts)
    return send_file(buf, as_attachment=True, download_name=f'{top}_attachments.zip',
                     mimetype='application/zip')

@app.route('/download/<step>/<session_id>')
def download(step, session_id):
    if session_id not in sessions: return jsonify({'error':'Session not found'}), 404
    sess = sessions[session_id]
    parts = sess['parts']; labor = sess['labor']; top = sess['top']
    seen, bom_rows = build_hierarchy(parts, top)

    used_colours = set()
    for p in parts:
        if 'Powder Coating' in p['processes'] and p['colour']:
            used_colours.add(p['colour'])

    # Build items
    item_data = [ITEM_COLS, irow(top, top+' Assembly', 'Make')]
    done = {top}
    for pn, p in seen.items():
        if pn in done: continue
        done.add(pn)
        existing_id, _ = get_existing_item(pn)
        if existing_id:
            continue
        item_data.append(irow(pn, p['desc'], p['origin'], p['mat'], p['thick']))
    for pc in sorted(used_colours):
        item_data.append(pc_irow(pc))

    # Build BOM
    part_procs = {p['pn']: p['processes'] for p in parts}
    bom_data = [BOM_H]
    for r in bom_rows:
        child_p = seen.get(r['child'])
        parent_procs = part_procs.get(r['parent'], [])
        child_origin = child_p['origin'] if child_p else 'Buy'
        routing_op = ''
        if child_origin == 'Buy' and 'Clinching' in parent_procs:
            routing_op = 'Clinching'
        bom_data.append([r['parent'],'',r['child'],'',r['qty'],'','','',routing_op])

    for p in parts:
        if 'Powder Coating' in p['processes'] and p['colour']:
            x = labor[p['pn']]['x']; y = labor[p['pn']]['y']
            powder_kg, _ = calc_powder(x, y)
            if powder_kg:
                bom_data.append([p['pn'], '', p['colour'], '', powder_kg, '', '', '', 'Powder Coating'])

    # Build Routing
    rout_data = [ROUT_H]
    for pn, p in seen.items():
        if p['origin'] == 'Buy' or not p['processes']: continue
        _, existing_ops = get_existing_item(pn)
        existing_ops_lower = [o.lower() for o in existing_ops]
        for i, op in enumerate(p['processes']):
            if op.lower() in existing_ops_lower:
                continue
            lt = calc_labor(pn, op, labor)
            rout_data.append([
                pn, '', op, i+1, '', '',
                'Fixed', SETUP_SECS, 'Second',
                'Fixed', lt if lt else '', 'Second' if lt else '',
                '', 'Fixed', '', '', '', '', '', ''
            ])

    if step == '1':
        wb = build_workbook(item_data, [BOM_H], [ROUT_H])
        fname = f'{top}_STEP1_Items.xlsx'
    else:
        wb = build_workbook([ITEM_COLS], bom_data, rout_data)
        fname = f'{top}_STEP2_BOM_Routing.xlsx'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/auto_attach/<session_id>', methods=['POST'])
def auto_attach(session_id):
    global item_id_cache
    if session_id not in sessions:
        return jsonify({'error': 'Session not found'}), 404
    if not FULCRUM_SESSION:
        return jsonify({'error': 'No FULCRUM_SESSION set — add it to your environment'}), 400

    sess = sessions[session_id]
    attachments = sess.get('attachments', {})
    parts = sess['parts']
    top = sess['top']

    if not attachments:
        return jsonify({'error': 'No attachments found in uploaded zip'}), 400

    data = request.get_json(silent=True) or {}
    assembly_fulcrum_id = data.get('assembly_id', '')
    if assembly_fulcrum_id:
        item_id_cache = build_item_id_cache(assembly_fulcrum_id)

    part_numbers = set(p['pn'].upper() for p in parts)
    part_numbers.add(top.upper())

    results = []
    attached = skipped = failed = 0

    for pn, files in attachments.items():
        matched_pn = None
        for p in sorted(part_numbers, key=len, reverse=True):
            if pn == p or pn.startswith(p) or p.startswith(pn):
                matched_pn = p
                break
        if not matched_pn:
            print(f'No match for {pn} in {part_numbers}')
            continue

        item_id, _ = get_existing_item(matched_pn)
        if not item_id:
            for filename, _ in files:
                results.append({'part': matched_pn, 'file': filename, 'success': False, 'error': 'Part not found in Fulcrum'})
                failed += 1
            continue

        for filename, file_bytes in files:
            try:
                ext = os.path.splitext(filename)[1].lower()
                mime = 'application/pdf' if ext == '.pdf' else 'application/octet-stream'
                upload_resp = requests.post(
                    f'{FULCRUM_BASE}/attachments',
                    headers=fulcrum_upload_headers(),
                    files={'File': (filename, file_bytes, mime)},
                    data={
                        'Detail.Owner.Type': 'item',
                        'Detail.Owner.Id': item_id,
                        'Detail.IsNoteAttachment': 'false',
                        'Detail.Description': filename,
                    },
                    timeout=30
                )
                if upload_resp.status_code in [200, 201]:
                    results.append({'part': matched_pn, 'file': filename, 'success': True})
                    attached += 1
                else:
                    results.append({'part': matched_pn, 'file': filename, 'success': False, 'error': f'HTTP {upload_resp.status_code}: {upload_resp.text[:100]}'})
                    failed += 1
            except Exception as e:
                results.append({'part': matched_pn, 'file': filename, 'success': False, 'error': str(e)})
                failed += 1

    return jsonify({'results': results, 'attached': attached, 'skipped': skipped, 'failed': failed})


if __name__ == '__main__':
    app.run(debug=True, port=5050)